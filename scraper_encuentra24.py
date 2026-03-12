#!/usr/bin/env python3
"""
RADAR DE LEADS - Expedia Motors (Consignación)
===============================================
Scraper de Encuentra24 Guatemala para detectar vehículos de agencia
en venta por particulares → leads de consignación.

Filtros:
- Año: 2018+
- Precio: hasta Q600,000
- Transmisión: Automática
- Kilometraje: hasta 100,000 km (solo km, NO millas)
- Solo vehículos de agencia (detección por keywords)
- Excluye lotes/dealers (solo particulares vendiendo)

Autor: Pablo Lara / Expedia Motors
"""

import requests
from bs4 import BeautifulSoup
import re
import json
import time
import random
from datetime import datetime, timedelta
from dataclasses import dataclass, field, asdict
from typing import Optional
import csv
import os

# ============================================================
# CONFIGURACIÓN - Ajustar según necesidad
# ============================================================

CONFIG = {
    # Filtros de búsqueda
    "year_min": 2018,
    "price_max_gtq": 600000,
    "km_max": 100000,
    "transmission": "automatic",
    
    # URL base con filtros aplicados en Encuentra24
    "base_url": "https://www.encuentra24.com/guatemala-en/cars-auto-trucks-used-car",
    
    # Paginación
    "max_pages": 35,  # Encuentra24 tiene ~33 páginas de usados
    "results_per_page": 30,
    
    # Rate limiting (ser respetuoso con el servidor)
    "delay_min_seconds": 2,
    "delay_max_seconds": 5,
    
    # Headers para simular navegador
    "headers": {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "es-GT,es;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
    },
    
    # Output
    "output_dir": "output",
    "seen_file": "seen_ads.json",  # Tracking de anuncios ya vistos
}

# Keywords que indican vehículo de agencia
AGENCY_KEYWORDS = [
    "de agencia", "agencia", "record de agencia", "récord de agencia",
    "mantenimiento en agencia", "mantenimientos en agencia",
    "servicio de agencia", "servicios de agencia",
    "comprado en agencia", "único dueño", "unico dueño",
    "un solo dueño", "un dueño", "primera dueña", "primer dueño",
    "siempre en agencia", "historial de agencia",
]

# Keywords que indican que NO es particular (es dealer/lote)
DEALER_KEYWORDS = [
    "lote de carros", "lote de vehiculos", "lote de vehículos",
    "financiamiento propio", "recibimos vehiculo",
    "recibimos vehículo", "aceptamos vehiculo",
    "visacuotas", "cuotas bac", "cuotas mastercard",
    "crédito bancario", "credito bancario",
    "varios disponibles", "tenemos disponible",
    "importadora", "automotriz", "motors sa",
]

# Keywords que indican millas (descalifica el anuncio)
MILES_KEYWORDS = [
    "millas", "miles ", "mi.", "mileage",
]


@dataclass
class VehicleLead:
    """Representa un lead de vehículo detectado."""
    id: str = ""
    title: str = ""
    make: str = ""
    model: str = ""
    year: int = 0
    price: float = 0.0
    price_currency: str = "GTQ"
    km: int = 0
    transmission: str = ""
    fuel: str = ""
    style: str = ""
    location: str = ""
    description: str = ""
    seller_name: str = ""
    seller_phone: str = ""
    url: str = ""
    is_agency: bool = False
    agency_signals: list = field(default_factory=list)
    is_dealer: bool = False
    dealer_signals: list = field(default_factory=list)
    has_miles: bool = False
    date_scraped: str = ""
    date_posted: str = ""
    score: int = 0  # Puntuación de calidad del lead (0-100)


class Encuentra24Scraper:
    """Scraper para Encuentra24 Guatemala - sección carros usados."""
    
    def __init__(self, config: dict = CONFIG):
        self.config = config
        self.session = requests.Session()
        self.session.headers.update(config["headers"])
        self.seen_ads = self._load_seen_ads()
        self.leads = []
        
    def _load_seen_ads(self) -> set:
        """Carga IDs de anuncios ya procesados para evitar duplicados."""
        seen_file = os.path.join(self.config["output_dir"], self.config["seen_file"])
        if os.path.exists(seen_file):
            with open(seen_file, "r") as f:
                data = json.load(f)
                return set(data.get("seen_ids", []))
        return set()
    
    def _save_seen_ads(self):
        """Guarda IDs de anuncios procesados."""
        os.makedirs(self.config["output_dir"], exist_ok=True)
        seen_file = os.path.join(self.config["output_dir"], self.config["seen_file"])
        with open(seen_file, "w") as f:
            json.dump({
                "seen_ids": list(self.seen_ads),
                "last_updated": datetime.now().isoformat(),
            }, f, indent=2)
    
    def _respectful_delay(self):
        """Delay aleatorio entre requests para no sobrecargar el servidor."""
        delay = random.uniform(
            self.config["delay_min_seconds"],
            self.config["delay_max_seconds"]
        )
        time.sleep(delay)
    
    def _fetch_page(self, url: str) -> Optional[BeautifulSoup]:
        """Descarga y parsea una página."""
        try:
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            return BeautifulSoup(response.text, "html.parser")
        except requests.RequestException as e:
            print(f"  ⚠️  Error fetching {url}: {e}")
            return None
    
    def _extract_listing_data(self, card_element) -> Optional[VehicleLead]:
        """
        Extrae datos de una card de listing en la página de resultados.
        
        Encuentra24 estructura sus cards con:
        - Link al detalle del anuncio
        - Nombre del vendedor y ubicación
        - Título/descripción del vehículo
        - Precio
        - Metadata: año, combustible, transmisión, km
        """
        lead = VehicleLead()
        lead.date_scraped = datetime.now().isoformat()
        
        # Extraer link al detalle
        link = card_element.find("a", href=True)
        if not link:
            return None
        
        href = link.get("href", "")
        if "/cars-auto-trucks-used-car/" not in href:
            return None
            
        lead.url = f"https://www.encuentra24.com{href}" if href.startswith("/") else href
        
        # Extraer ID del anuncio desde la URL
        id_match = re.search(r"/(\d+)$", href)
        if id_match:
            lead.id = id_match.group(1)
        
        # Extraer texto completo de la card para análisis
        card_text = card_element.get_text(separator=" ", strip=True).lower()
        
        # Extraer vendedor y ubicación
        # Estructura típica: "NombreVendedor\nUbicación\nMarca Modelo\nDescripción..."
        text_parts = card_element.get_text(separator="\n", strip=True).split("\n")
        text_parts = [p.strip() for p in text_parts if p.strip()]
        
        # Buscar nombre del vendedor (primer texto significativo)
        for part in text_parts:
            if len(part) > 2 and not part.startswith("Q") and not part.startswith("$"):
                lead.seller_name = part
                break
        
        # Buscar ubicación (zona, Guatemala City, etc.)
        for part in text_parts:
            part_lower = part.lower()
            if any(loc in part_lower for loc in ["zona", "guatemala", "mixco", "villa nueva", 
                                                   "carretera", "antigua", "escuintla"]):
                lead.location = part
                break
        
        # Extraer precio
        price_patterns = [
            r'Q\s*(\d[\d,\.]+)',       # Q215,000 o Q 215000
            r'\$\s*(\d[\d,\.]+)',       # $58,000
        ]
        for pattern in price_patterns:
            price_match = re.search(pattern, card_element.get_text())
            if price_match:
                price_str = price_match.group(1).replace(",", "").replace(".", "")
                try:
                    lead.price = float(price_str)
                    lead.price_currency = "USD" if "$" in pattern else "GTQ"
                except ValueError:
                    pass
                break
        
        # Extraer metadata (año, combustible, transmisión, km)
        # Estos vienen en formato: "2022 · Gasoline · Automatic · 46000 km"
        metadata_items = re.findall(
            r'(\d{4})\s*[·•]\s*(Gasoline|Diesel|Electric|Híbrido|Gas)\s*[·•]\s*'
            r'(Automatic|Manual|5\+)\s*[·•]\s*(\d[\d,]*)\s*km',
            card_element.get_text(), re.IGNORECASE
        )
        
        if metadata_items:
            year_str, fuel, trans, km_str = metadata_items[0]
            lead.year = int(year_str)
            lead.fuel = fuel
            lead.transmission = trans
            lead.km = int(km_str.replace(",", ""))
        else:
            # Intentar extraer campos individualmente
            year_match = re.search(r'\b(20[12]\d)\b', card_text)
            if year_match:
                lead.year = int(year_match.group(1))
            
            km_match = re.search(r'(\d[\d,]*)\s*km', card_text)
            if km_match:
                lead.km = int(km_match.group(1).replace(",", ""))
            
            if "automatic" in card_text or "automátic" in card_text:
                lead.transmission = "Automatic"
            elif "manual" in card_text or "mecánic" in card_text:
                lead.transmission = "Manual"
        
        # Extraer marca y modelo del título
        # Estructura típica del link text: "Marca Modelo Descripción"
        title_elem = card_element.find(["h2", "h3", "strong"])
        if title_elem:
            lead.title = title_elem.get_text(strip=True)
        else:
            # Buscar en el texto de la card
            for part in text_parts:
                if any(make in part for make in ["Toyota", "Honda", "Mazda", "Kia", "Hyundai",
                                                   "BMW", "Mercedes", "Audi", "Volkswagen",
                                                   "Nissan", "Mitsubishi", "Chevrolet", "Ford",
                                                   "Jeep", "Land Rover", "Porsche", "Lexus",
                                                   "Subaru", "Suzuki", "Volvo", "MINI",
                                                   "Changan", "Renault", "Dodge", "Peugeot"]):
                    lead.title = part
                    break
        
        # Extraer descripción
        desc_parts = [p for p in text_parts if len(p) > 50]
        if desc_parts:
            lead.description = desc_parts[0][:500]
        
        return lead
    
    def _check_agency_signals(self, lead: VehicleLead, full_text: str) -> VehicleLead:
        """Detecta señales de que el vehículo es de agencia."""
        text_lower = full_text.lower()
        
        # Buscar keywords de agencia
        for keyword in AGENCY_KEYWORDS:
            if keyword in text_lower:
                lead.is_agency = True
                lead.agency_signals.append(keyword)
        
        # Buscar keywords de dealer
        for keyword in DEALER_KEYWORDS:
            if keyword in text_lower:
                lead.is_dealer = True
                lead.dealer_signals.append(keyword)
        
        # Detectar millas
        for keyword in MILES_KEYWORDS:
            if keyword in text_lower:
                lead.has_miles = True
                break
        
        return lead
    
    def _score_lead(self, lead: VehicleLead) -> int:
        """
        Calcula un score de calidad del lead (0-100).
        Más alto = más probable que sea un buen lead de consignación.
        """
        score = 50  # Base score
        
        # Señales positivas
        if lead.is_agency:
            score += 20
        if "único dueño" in " ".join(lead.agency_signals).lower():
            score += 10
        if "record de agencia" in " ".join(lead.agency_signals).lower():
            score += 10
        if lead.km > 0 and lead.km < 30000:
            score += 10
        elif lead.km > 0 and lead.km < 50000:
            score += 5
        if lead.year >= 2022:
            score += 5
        if lead.seller_name and len(lead.seller_name.split()) >= 2:
            score += 5  # Nombre completo = más probable particular
        
        # Señales negativas
        if lead.is_dealer:
            score -= 30
        if lead.has_miles:
            score -= 50  # Descalificante
        if lead.transmission.lower() == "manual":
            score -= 50  # Solo automáticos
        if lead.year < 2018:
            score -= 50  # Fuera de rango
        
        # Precio en USD: convertir aproximado (1 USD ≈ 7.8 GTQ)
        price_gtq = lead.price
        if lead.price_currency == "USD":
            price_gtq = lead.price * 7.8
        if price_gtq > 600000:
            score -= 30
        
        return max(0, min(100, score))
    
    def _passes_filters(self, lead: VehicleLead) -> bool:
        """Verifica si el lead pasa los filtros configurados."""
        # Filtro de año
        if lead.year > 0 and lead.year < self.config["year_min"]:
            return False
        
        # Filtro de precio
        price_gtq = lead.price
        if lead.price_currency == "USD":
            price_gtq = lead.price * 7.8
        if price_gtq > self.config["price_max_gtq"] and lead.price > 0:
            return False
        
        # Filtro de transmisión (solo automáticos)
        if lead.transmission and lead.transmission.lower() in ["manual", "mecánic", "mecanico"]:
            return False
        
        # Filtro de kilometraje
        if lead.km > self.config["km_max"] and lead.km > 0:
            return False
        
        # Filtro de millas (descalificante)
        if lead.has_miles:
            return False
        
        return True
    
    def scrape_listing_page(self, page_num: int = 1) -> list[VehicleLead]:
        """Scrapea una página de listings."""
        if page_num == 1:
            url = self.config["base_url"]
        else:
            url = f"{self.config['base_url']}.{page_num}"
        
        print(f"  📄 Scraping página {page_num}: {url}")
        
        soup = self._fetch_page(url)
        if not soup:
            return []
        
        leads = []
        
        # Encuentra24 usa diferentes estructuras de cards
        # Buscar todos los links que apuntan a anuncios individuales
        ad_links = soup.find_all("a", href=re.compile(r"/cars-auto-trucks-used-car/.+/\d+$"))
        
        # También buscar por estructura de cards
        cards = soup.find_all(["div", "article"], class_=re.compile(r"ann-box|listing|result|card", re.I))
        
        # Si no encontramos cards con clase, usar los links como base
        elements = cards if cards else []
        
        # Procesar links únicos
        seen_urls = set()
        for link in ad_links:
            href = link.get("href", "")
            if href in seen_urls:
                continue
            seen_urls.add(href)
            
            # Buscar el contenedor padre de la card
            parent = link
            for _ in range(5):
                if parent.parent:
                    parent = parent.parent
                    parent_text = parent.get_text(strip=True)
                    if len(parent_text) > 100:  # Encontramos la card completa
                        break
            
            lead = self._extract_listing_data(parent)
            if lead and lead.id:
                # Verificar señales en el texto de la card
                card_text = parent.get_text(separator=" ", strip=True)
                lead = self._check_agency_signals(lead, card_text)
                leads.append(lead)
        
        print(f"    → Encontrados {len(leads)} anuncios en página {page_num}")
        return leads
    
    def scrape_detail_page(self, lead: VehicleLead) -> VehicleLead:
        """
        Scrapea la página de detalle de un anuncio para obtener
        descripción completa, teléfono y más señales.
        """
        print(f"    🔍 Detalle: {lead.title[:50]}...")
        
        soup = self._fetch_page(lead.url)
        if not soup:
            return lead
        
        full_text = soup.get_text(separator=" ", strip=True)
        
        # Re-chequear señales con el texto completo
        lead = self._check_agency_signals(lead, full_text)
        
        # Extraer descripción completa
        desc_elem = soup.find(["div", "p"], class_=re.compile(r"desc|detail|content", re.I))
        if desc_elem:
            lead.description = desc_elem.get_text(strip=True)[:1000]
        
        # Extraer teléfono (si está visible)
        phone_match = re.search(r'(\+?502[\s-]?\d{4}[\s-]?\d{4}|\d{4}[\s-]?\d{4})', full_text)
        if phone_match:
            lead.seller_phone = phone_match.group(1).strip()
        
        # Extraer estilo de vehículo
        for style in ["SUV", "Sedan", "Pickup", "Hatchback", "Coupe", "Convertible", "Minivan"]:
            if style.lower() in full_text.lower():
                lead.style = style
                break
        
        # Verificar millas en texto completo
        if not lead.has_miles:
            text_lower = full_text.lower()
            for kw in MILES_KEYWORDS:
                if kw in text_lower:
                    # Verificar que no sea un falso positivo (ej: "Milwaukee")
                    context = text_lower[max(0, text_lower.index(kw)-30):text_lower.index(kw)+30]
                    if any(num in context for num in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]):
                        lead.has_miles = True
                        break
        
        return lead
    
    def run(self, full_detail: bool = False, only_new: bool = True):
        """
        Ejecuta el scraping completo.
        
        Args:
            full_detail: Si True, visita cada anuncio para obtener más datos.
                        Más lento pero más preciso.
            only_new: Si True, solo procesa anuncios no vistos anteriormente.
        """
        print("=" * 60)
        print("🚗 RADAR DE LEADS - Expedia Motors (Consignación)")
        print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print("=" * 60)
        print(f"\nFiltros: Año ≥{self.config['year_min']} | "
              f"Precio ≤Q{self.config['price_max_gtq']:,.0f} | "
              f"KM ≤{self.config['km_max']:,} | Automático | Solo KM")
        print()
        
        all_leads = []
        
        # Fase 1: Scrapear páginas de listings
        print("📡 FASE 1: Scraping de listings...")
        for page in range(1, self.config["max_pages"] + 1):
            page_leads = self.scrape_listing_page(page)
            
            if not page_leads:
                print(f"    ⏹  No más resultados en página {page}. Deteniendo.")
                break
            
            all_leads.extend(page_leads)
            self._respectful_delay()
        
        print(f"\n📊 Total anuncios encontrados: {len(all_leads)}")
        
        # Filtrar anuncios ya vistos
        if only_new:
            new_leads = [l for l in all_leads if l.id not in self.seen_ads]
            print(f"🆕 Anuncios nuevos: {len(new_leads)}")
            all_leads = new_leads
        
        # Fase 2: Aplicar filtros básicos
        print("\n🔧 FASE 2: Aplicando filtros...")
        filtered_leads = [l for l in all_leads if self._passes_filters(l)]
        print(f"✅ Pasan filtros: {len(filtered_leads)}")
        
        # Fase 3 (opcional): Obtener detalle de cada anuncio
        if full_detail and filtered_leads:
            print(f"\n🔬 FASE 3: Obteniendo detalles ({len(filtered_leads)} anuncios)...")
            for i, lead in enumerate(filtered_leads):
                lead = self.scrape_detail_page(lead)
                self._respectful_delay()
                
                # Re-aplicar filtros con datos más completos
                if not self._passes_filters(lead):
                    filtered_leads[i] = None
            
            filtered_leads = [l for l in filtered_leads if l is not None]
            print(f"✅ Pasan filtros post-detalle: {len(filtered_leads)}")
        
        # Fase 4: Scoring y ordenamiento
        print("\n📈 FASE 4: Scoring de leads...")
        for lead in filtered_leads:
            lead.score = self._score_lead(lead)
        
        # Ordenar por score (mejor primero)
        filtered_leads.sort(key=lambda l: l.score, reverse=True)
        
        # Guardar IDs como vistos
        for lead in all_leads:
            if lead.id:
                self.seen_ads.add(lead.id)
        self._save_seen_ads()
        
        self.leads = filtered_leads
        
        # Resumen
        print(f"\n{'=' * 60}")
        print(f"🏆 RESULTADOS: {len(filtered_leads)} leads calificados")
        if filtered_leads:
            agency_leads = [l for l in filtered_leads if l.is_agency]
            high_score = [l for l in filtered_leads if l.score >= 70]
            print(f"   🏪 De agencia confirmado: {len(agency_leads)}")
            print(f"   ⭐ Score alto (≥70): {len(high_score)}")
        print("=" * 60)
        
        return filtered_leads
    
    def export_to_excel(self, leads: list[VehicleLead] = None, filename: str = None):
        """Exporta los leads a un archivo Excel (.xlsx)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        if leads is None:
            leads = self.leads
        
        if not leads:
            print("⚠️  No hay leads para exportar.")
            return None
        
        if filename is None:
            date_str = datetime.now().strftime("%Y-%m-%d")
            filename = f"leads_consignacion_{date_str}.xlsx"
        
        filepath = os.path.join(self.config["output_dir"], filename)
        os.makedirs(self.config["output_dir"], exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Consignación"
        
        # Estilos
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        high_score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        medium_score_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        agency_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Headers
        headers = [
            "Score", "De Agencia", "Marca/Modelo", "Año", "Precio",
            "Moneda", "KM", "Transmisión", "Combustible", "Estilo",
            "Vendedor", "Teléfono", "Ubicación", "Señales Agencia",
            "URL", "Descripción"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Datos
        for row_idx, lead in enumerate(leads, 2):
            data = [
                lead.score,
                "✅ SÍ" if lead.is_agency else "❌ No",
                lead.title or f"{lead.make} {lead.model}",
                lead.year if lead.year > 0 else "",
                lead.price if lead.price > 0 else "",
                lead.price_currency,
                lead.km if lead.km > 0 else "",
                lead.transmission,
                lead.fuel,
                lead.style,
                lead.seller_name,
                lead.seller_phone,
                lead.location,
                ", ".join(lead.agency_signals) if lead.agency_signals else "",
                lead.url,
                lead.description[:200] if lead.description else "",
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(col >= 14))
                
                # Color por score
                if col == 1:
                    if lead.score >= 70:
                        cell.fill = high_score_fill
                    elif lead.score >= 50:
                        cell.fill = medium_score_fill
                
                # Color si es de agencia
                if col == 2 and lead.is_agency:
                    cell.fill = agency_fill
        
        # Ajustar anchos de columna
        column_widths = [8, 12, 30, 8, 12, 8, 10, 14, 12, 12, 20, 15, 18, 25, 50, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Congelar header
        ws.freeze_panes = "A2"
        
        # Auto-filtro
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(leads) + 1}"
        
        # Hoja de resumen
        ws_summary = wb.create_sheet("Resumen")
        summary_data = [
            ["RADAR DE LEADS - Expedia Motors", ""],
            ["Fecha", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["", ""],
            ["Total leads calificados", len(leads)],
            ["De agencia confirmado", len([l for l in leads if l.is_agency])],
            ["Score alto (≥70)", len([l for l in leads if l.score >= 70])],
            ["Score medio (50-69)", len([l for l in leads if 50 <= l.score < 70])],
            ["Score bajo (<50)", len([l for l in leads if l.score < 50])],
            ["", ""],
            ["FILTROS APLICADOS", ""],
            ["Año mínimo", self.config["year_min"]],
            ["Precio máximo", f"Q{self.config['price_max_gtq']:,.0f}"],
            ["KM máximo", f"{self.config['km_max']:,}"],
            ["Transmisión", "Solo automáticos"],
            ["Kilometraje", "Solo KM (excluye millas)"],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=(row_idx in [1, 10]))
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 20
        
        wb.save(filepath)
        print(f"\n📁 Excel exportado: {filepath}")
        return filepath
    
    def export_to_csv(self, leads: list[VehicleLead] = None, filename: str = None):
        """Exporta los leads a CSV como respaldo."""
        if leads is None:
            leads = self.leads
        
        if not leads:
            return None
        
        if filename is None:
            date_str = datetime.now().strftime("%Y-%m-%d")
            filename = f"leads_consignacion_{date_str}.csv"
        
        filepath = os.path.join(self.config["output_dir"], filename)
        os.makedirs(self.config["output_dir"], exist_ok=True)
        
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "Score", "De Agencia", "Título", "Año", "Precio", "Moneda",
                "KM", "Transmisión", "Combustible", "Vendedor", "Teléfono",
                "Ubicación", "Señales Agencia", "URL"
            ])
            for lead in leads:
                writer.writerow([
                    lead.score,
                    "SÍ" if lead.is_agency else "No",
                    lead.title,
                    lead.year,
                    lead.price,
                    lead.price_currency,
                    lead.km,
                    lead.transmission,
                    lead.fuel,
                    lead.seller_name,
                    lead.seller_phone,
                    lead.location,
                    "; ".join(lead.agency_signals),
                    lead.url,
                ])
        
        print(f"📁 CSV exportado: {filepath}")
        return filepath


def main():
    """Punto de entrada principal."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Radar de Leads - Expedia Motors")
    parser.add_argument("--full-detail", action="store_true",
                       help="Visitar cada anuncio para obtener datos completos (más lento)")
    parser.add_argument("--all", action="store_true",
                       help="Procesar todos los anuncios, no solo nuevos")
    parser.add_argument("--max-pages", type=int, default=35,
                       help="Máximo de páginas a scrapear (default: 35)")
    parser.add_argument("--output", type=str, default=None,
                       help="Nombre del archivo de salida")
    
    args = parser.parse_args()
    
    CONFIG["max_pages"] = args.max_pages
    
    scraper = Encuentra24Scraper(CONFIG)
    leads = scraper.run(
        full_detail=args.full_detail,
        only_new=not args.all,
    )
    
    if leads:
        scraper.export_to_excel(leads, args.output)
        scraper.export_to_csv(leads)
    else:
        print("\n😕 No se encontraron leads nuevos que pasen los filtros.")
        print("   Intenta con --all para ver todos los anuncios.")


if __name__ == "__main__":
    main()
